from django.shortcuts import render,redirect, get_object_or_404
from performanceReportTracking.models import Staff,performanceReport,Customer,customerCatogery,customerPackage,Customer,customerPlatform,Admin
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
import os
from django.contrib import messages
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from collections import defaultdict


#LOGIN PAGE
def login(request):
    if request.method == 'POST':
        
        # STAFF LOGIN
        if 'staffID' in request.POST:
            staffID = request.POST['staffID']
            staffPassword = request.POST['staffPass']

            if staffID == "admin" and staffPassword == "admin":
                request.session['staffID'] = staffID
                request.session['staffName'] = "Admin"  # You can replace this with actual admin name if applicable
                return redirect('home')

            try:
                staff = Staff.objects.get(staffID=staffID)
                if staff.staffPassword == staffPassword: 
                    request.session['staffID'] = staffID
                    request.session['staffName'] = staff.staffName  # Store the staff name in the session
                    return redirect('home')
                else:
                    return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})

            except Staff.DoesNotExist:
                return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})

        # ADMIN LOGIN
        elif 'adminID' in request.POST:
            adminID = request.POST['adminID']
            adminPassword = request.POST['adminPass']
            
            if adminID == "admin" and adminPassword == "admin":
                request.session['adminID'] = adminID
                request.session['adminName'] = "Admin"  # You can replace this with actual admin name if applicable
                return redirect('adminHome')

            try:
                admin = Admin.objects.get(adminID=adminID)
                if admin.adminPassword == adminPassword: 
                    request.session['adminID'] = adminID
                    request.session['adminName'] = admin.adminName  # Store the admin name in the session
                    return redirect('adminhome')
                else:
                    return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})

            except Admin.DoesNotExist:
                return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})

    return render(request, 'login.html')

#HOME PAGE
def home(request):
    submitted_reports_count = performanceReport.objects.count()
    
    staff_name = request.session.get('staffName')

    context = {
        'submitted_reports_count': submitted_reports_count,
    }
    return render(request, "home.html", context)

#ADMIN HOME PAGE
def adminhome(request):
    submitted_reports_count = performanceReport.objects.count()

    context = {
        'submitted_reports_count': submitted_reports_count,
    }
    return render(request, "adminhome.html", context)

#MANAGE STAFF PAGE
def managestaff(request):
    if request.method == "POST":
        staff_id = request.POST.get('staffID')  # Changed from 'staffID'
        staff_name = request.POST.get('staffName')
        staff_password = request.POST.get('staffPassword')
        staff_email = request.POST.get('staffEmail')

        if staff_id and staff_name and staff_password and staff_email:
            try:
                new_staff = Staff(
                    staffID=staff_id,
                    staffName=staff_name,
                    staffPassword=staff_password,
                    staffEmail=staff_email
                )
                new_staff.save()
                messages.success(request, 'Staff added successfully.')
                return redirect('managestaff')  

            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
        else:
            messages.error(request, 'Please provide all required fields.')

    return render(request, 'managestaff.html') 

#DELETE STAFF ADMIN PAGE
def deletestaffadmin(request):
    query = request.GET.get('query', '')  
    if query:
        list_staff = Staff.objects.filter(staffID__icontains=query)
    else:
        list_staff = Staff.objects.all()

    return render(request, 'deletestaffadmin.html', {
        'list_staff': list_staff,
        'query': query,
    })
    
#DELETE STAFF
def deletestaff(request, staff_id):
    if request.method == 'POST':
        try:
            staff_member = Staff.objects.get(staffID=staff_id)  
            staff_member.delete()
            return JsonResponse({'success': True})
        except Staff.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Staff member not found.'})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

#VIEW SUBMISSION ADMIN PAGE
def viewsubmissionadmin(request):
    query = request.GET.get('query', '')  
    month = request.GET.get('month', None) 
    year = request.GET.get('year', None) 
    selected_report_id = request.GET.get('report_id')

    filters = Q()
    
    if selected_report_id:
        filters &= Q(id=selected_report_id)

    if query:
        filters &= Q(companyName__exact=query)

    if month and year:
        try:
            month = int(month)
            year = int(year)
            print(f"Filtering by month: {month}, year: {year}")
            filters &= Q(dateSubmission__month=month) & Q(dateSubmission__year=year)
        except ValueError:
            print("Invalid month or year provided")  
            pass

    list_customer = performanceReport.objects.filter(filters)

    print(f"Found {list_customer.count()} results")

    context = {
        'list_customer': list_customer,
        'query': query,
        'month': month,
        'year': year,
        'selected_report_id': selected_report_id,
    }
    return render(request, "viewsubmissionadmin.html", context)

#DETAIL SUBMISSION ADMIN PAGE
def detailsubmissionadmin(request, report_id):
    report = get_object_or_404(performanceReport, id=report_id)

    reportFileUrl = report.report.url if report.report else None
    reportFileName = report.report.name.split('/')[-1] if report.report else None

    context = {
        'report_id': report_id,  
        'companyName': report.companyName.companyName,
        'companyEmail': report.companyName.customerEmail,
        'catogery': report.companyName.catogery.katogeri,
        'package': report.companyName.package.pakej,
        'adsPlatform': report.companyName.adsPlatform.ads,
        'staffID': report.staffID.staffID,
        'reportFileUrl': reportFileUrl,
        'reportFileName': reportFileName, 
        'duration': report.duration,
        'dateSubmission': report.dateSubmission.strftime('%d-%m-%Y') if report.dateSubmission else '--',
        'remark': report.remark,
    }
    
    return render(request, "detailsubmissionadmin.html", context)

#ADMIN PROFILE PAGE
def adminprofile(request):
    adminID = request.session.get('adminID')

    if not adminID:
        print("adminID not found in session")
        return HttpResponse("Admin not logged in")
    
    try:
        admin = Admin.objects.get(adminID=adminID)
        print("Admin found:", admin)

        context = {
            'adminID': admin.adminID,
            'adminName': admin.adminName,
            'adminEmail': admin.adminEmail,
        }
        return render(request, "adminprofile.html", context)
    except Admin.DoesNotExist:
        print("Admin not found for adminID:", adminID)
        return HttpResponse("Admin not found")



#CUSTOMER SUBMISSION PAGE
def customersubmission (request):
    if request.method == "POST":
        company_name = request.POST.get('companyName')
        customer_email = request.POST.get('companyEmail')
        category_name = request.POST.get('catogery')
        package_name = request.POST.get('package')
        ads_platform_name = request.POST.get('adsPlatform')

        if company_name and customer_email and category_name and package_name and ads_platform_name:
            try:
                category = customerCatogery.objects.get(katogeri=category_name)
                package = customerPackage.objects.get(pakej=package_name)
                ads_platform = customerPlatform.objects.get(ads=ads_platform_name)

                new_customer = Customer(
                    companyName=company_name,
                    customerEmail=customer_email,
                    catogery=category,
                    package=package,
                    adsPlatform=ads_platform
                )
                new_customer.save()
                messages.success(request, 'Customer added successfully.')
                return redirect('customersubmission') 

            except customerCatogery.DoesNotExist:
                messages.error(request, 'Category does not exist.')
            except customerPackage.DoesNotExist:
                messages.error(request, 'Package does not exist.')
            except customerPlatform.DoesNotExist:
                messages.error(request, 'Ads Platform does not exist.')
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
        else:
            messages.error(request, 'Please provide all required fields.')

    return render(request, 'customersubmission.html', {
        'categories': customerCatogery.objects.all(),
        'packages': customerPackage.objects.all(),
        'ads_platforms': customerPlatform.objects.all(),
    })

#REPORT SUBMISSION PAGE
def reportsubmission(request):
    staffID = request.session.get('staffID')
    
    if not staffID:
        return render(request, 'error.html', {'message': 'Staff ID not found in session'})

    staff = get_object_or_404(Staff, staffID=staffID) 
    
    if request.method == "POST":
        company_name = request.POST.get('companyName')
        date_submission = request.POST.get('dateSubmission')
        date_duration_start = request.POST.get('dateDurationStart')
        date_duration_end = request.POST.get('dateDurationEnd')
        remark = request.POST.get('remark')
        report_file = request.FILES.get('report')

        if date_duration_start and date_duration_end:
            duration = f"{date_duration_start} to {date_duration_end}"
        else:
            duration = None

        print(f"Company Name: {company_name}")
        print(f"Report File: {report_file}")
        print(f"Date Submission: {date_submission}")
        print(f"Duration: {duration}")
        print(f"Remark: {remark}")

        if not company_name:
            messages.error(request, "Company name is required.")
        if not report_file:
            messages.error(request, "Report file is required.")
        if not duration:
            messages.error(request, "Duration is required.")
        if not date_submission:
            messages.error(request, "Date of submission is required.")
        if not remark:
            messages.error(request, "Remark is required.")

        if company_name and report_file and duration and date_submission and remark:
            try:
                customer = get_object_or_404(Customer, companyName=company_name)
                
                new_performance_report = performanceReport(
                    staffID=staff,
                    companyName=customer,
                    duration=duration,
                    remark=remark,
                    dateSubmission=date_submission,
                    report=report_file
                )
                new_performance_report.save()

                messages.success(request, "Performance report submitted successfully.")
                return redirect('reportsubmission') 
            except Customer.DoesNotExist:
                messages.error(request, 'Customer does not exist.')
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
        else:
            messages.error(request, "Please fill in all required fields.")
    
    context = {
        'staffID': staffID,
        'staffs': Staff.objects.all() 
    }
    return render(request, 'reportsubmission.html', context)
       
#VIEW SUBMISSION PAGE
def viewsubmission(request):
    query = request.GET.get('query', '')  
    month = request.GET.get('month', None) 
    year = request.GET.get('year', None) 
    selected_report_id = request.GET.get('report_id')

    filters = Q()
    
    if selected_report_id:
        filters &= Q(id=selected_report_id)

    if query:
        filters &= Q(companyName__exact=query)

    if month and year:
        try:
            month = int(month)
            year = int(year)
            print(f"Filtering by month: {month}, year: {year}")
            filters &= Q(dateSubmission__month=month) & Q(dateSubmission__year=year)
        except ValueError:
            print("Invalid month or year provided")  
            pass

    list_customer = performanceReport.objects.filter(filters)

    print(f"Found {list_customer.count()} results")

    context = {
        'list_customer': list_customer,
        'query': query,
        'month': month,
        'year': year,
        'selected_report_id': selected_report_id,
    }
    return render(request, 'viewsubmission.html', context)

#DELETE SUBMISSION 
def deleteReport(request, report_id):
    try:
        report = get_object_or_404(performanceReport, pk=report_id)
        report.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

#DETAIL SUBMISSION PAGE
def detailsubmission(request, report_id):
    report = get_object_or_404(performanceReport, id=report_id)

    reportFileUrl = report.report.url if report.report else None
    reportFileName = report.report.name.split('/')[-1] if report.report else None

    context = {
        'report_id': report_id,  
        'companyName': report.companyName.companyName,
        'companyEmail': report.companyName.customerEmail,
        'catogery': report.companyName.catogery.katogeri,
        'package': report.companyName.package.pakej,
        'adsPlatform': report.companyName.adsPlatform.ads,
        'staffID': report.staffID.staffID,
        'reportFileUrl': reportFileUrl,
        'reportFileName': reportFileName, 
        'duration': report.duration,
        'dateSubmission': report.dateSubmission.strftime('%d-%m-%Y') if report.dateSubmission else '--',
        'remark': report.remark,
    }
    
    return render(request, 'detailsubmission.html', context)

#DOWNLOAD REPORT
def download_report(request, report_id):
    report = get_object_or_404(performanceReport, id=report_id)
    file_path = report.report.path

    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.ms-powerpoint')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response
    else:
        return HttpResponse("File not found", status=404)

#EDIT SUBMISSION PAGE
def editsubmission(request, report_id):
    report = get_object_or_404(performanceReport, id=report_id)

    if request.method == 'POST':
        print("POST Data:", request.POST)
        print("FILES Data:", request.FILES)

        report.duration = request.POST.get('duration')
        report.dateSubmission = request.POST.get('dateSubmission')
        report.remark = request.POST.get('remark')

        if 'report' in request.FILES:
            report.report = request.FILES['report']
        
        try:
            report.save()
            print("Report updated successfully:", report)
            return redirect('detailsubmission', report_id=report.id)
        except Exception as e:
            print("Error saving report:", e)
            return HttpResponse("An error occurred while saving the report.", status=500)

    context = {
        'report': report,
        'companyName': report.companyName.companyName,
        'companyEmail': report.companyName.customerEmail,
        'catogery': report.companyName.catogery.katogeri,
        'package': report.companyName.package.pakej,
        'adsPlatform': report.companyName.adsPlatform.ads,
        'staffID': report.staffID.staffID,
        'reportFileUrl': report.report.url if report.report else None,
        'reportFileName': report.report.name if report.report else None,
        'duration': report.duration,
        'dateSubmission': report.dateSubmission.strftime('%d-%m-%Y') if report.dateSubmission else '--',
        'remark': report.remark,
    }

    return render(request, 'editsubmission.html', context)

def save_updateReport(request, report_id):
    if request.method == 'POST':
        report = get_object_or_404(performanceReport, id=report_id)
        
        report.duration = request.POST.get('duration')
        report.dateSubmission = request.POST.get('dateSubmission')
        report.remark = request.POST.get('remark')

        if 'report' in request.FILES:
            report.report = request.FILES['report']

        report.save()

        return redirect('detailsubmission', report_id=report.id)

    return redirect('detailsubmission', report_id=report_id)


#PROFILE PAGE
def profile(request):
    staffID = request.session.get('staffID')

    if not staffID:
        print("staffID not found in session")
        return HttpResponse("Admin not logged in")
    
    try:
        staff = Staff.objects.get(staffID=staffID)
        print("Staff found:", staff)

        context = {
            'staffID': staff.staffID,
            'staffName': staff.staffName,
            'staffEmail': staff.staffEmail,
        }
        return render(request, "profile.html", context)
    except Staff.DoesNotExist:
        print("Staff not found for staffID:", staffID)
        return HttpResponse("Staff not found")
    
#EXPORT TO EXCEL
def performance_report_excel(request):
    qs = performanceReport.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Performance_Report.xlsx"'
    
    workbook = Workbook()

    data_by_month = defaultdict(list)
    for report in qs:
        month_year = report.dateSubmission.strftime('%B %Y')  
        data_by_month[month_year].append(report)

    for month, reports in data_by_month.items():
        worksheet = workbook.create_sheet(title=month[:31])  

        worksheet.merge_cells('A1:F1')
        worksheet['A1'] = f"Performance Report - {month}"
        worksheet['A1'].fill = PatternFill(start_color="246ba1", end_color="246ba1", fill_type="solid")
        worksheet['A1'].font = Font(bold=True, color="F7F6FA")
        worksheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

        columns = ['Staff ID', 'Company Name', 'Company Email', 'Duration', 'Date Submission', 'Remark']
        row_num = 2 
        
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")
            cell.font = Font(bold=True, color="F7F6FA")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for report in reports:
            row_num += 1
            row = [
                report.staffID.staffID,  
                report.companyName.companyName,
                report.companyName.customerEmail,
                report.duration,
                report.dateSubmission.strftime('%d-%m-%Y') if report.dateSubmission else '--',
                report.remark
            ]
            
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    workbook.save(response)
    return response